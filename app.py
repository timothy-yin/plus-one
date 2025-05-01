
# app3.py
import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
import tempfile
import openpyxl
from openpyxl.styles import PatternFill, Font
import datetime  # 放在 import 區塊（如果還沒加）

st.set_page_config(page_title="PLUS_ONE", layout="wide")
st.title("📚 PLOS ONE 期刊作者與機構擷取工具")

# ✅ 一定放在這裡才會一打開就顯示
st.info(f"🔧 測試版本 v1.11 ‧ 載入時間：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
doi_input = st.text_area("請輸入 DOI（每行一筆，限PLOS ONE）")
run_button = st.button("🚀 開始擷取")

if run_button and doi_input.strip():
    st.markdown("🏃‍♂️ 正在擷取中，請稍候...")

    with st.spinner("資料擷取中..."):
        dois = [d.strip() for d in doi_input.strip().split("\n")]
        all_records = []
        seen_names = set()  # 用來記錄已經出現過的作者姓名，避免重複

        for doi in dois:
            url = f"https://journals.plos.org/plosone/article/file?id={doi}&type=manuscript"
            res = requests.get(url)

            if res.status_code == 200:
               soup = BeautifulSoup(res.content, "lxml-xml")

               # 文章標題
               article_title = soup.find("article-title")
               title = article_title.get_text(strip=True) if article_title else "N/A"

               # 建立 aff 對應表
               aff_dict = {
                   aff.get("id"): aff.find("addr-line").get_text(strip=True) if aff.find("addr-line") else aff.get_text(strip=True)
                   for aff in soup.find_all("aff")
               }

               # 🔍 只在 <front> 的 <contrib-group> 裡抓作者，避免抓到附錄或參與名單
               author_group = soup.find("contrib-group")
               authors = author_group.find_all("contrib", {"contrib-type": "author"}) if author_group else []


               for idx, author in enumerate(authors, start=1):
                   surname = author.find("surname")
                   given_names = author.find("given-names")
                   name = f"{given_names.text.strip()} {surname.text.strip()}" if given_names and surname else surname.text.strip() if surname else "N/A"

                # 避免重複作者名
                   if name in seen_names:
                       continue
                   seen_names.add(name)

                   # 標記第一作者 / 通訊作者
                   if idx == 1:
                      name += "（第一作者）"
                   elif author.find("xref", {"ref-type": "corresp"}):
                      name += "（通訊作者）"

                   # 找對應機構
                   aff_ref = author.find("xref", {"ref-type": "aff"})
                   aff_id = aff_ref.get("rid") if aff_ref else None
                   affiliation = aff_dict.get(aff_id, "N/A")

                   # 儲存資料
                   all_records.append({
                       "Title": title,
                       "Name": name,
                       "Order": idx,
                       "Affiliation": affiliation,
                       "DOI": doi
                   })

            else:
                st.warning(f"❌ 無法取得 DOI: {doi}")


    df = pd.DataFrame(all_records)
    df["IsFirst"] = df["Order"] == 1

    if not df.empty:
        st.success("✅ 擷取成功！")
        st.subheader("📋 擷取結果")

        # ✅ 畫面用 HTML 呈現樣式
        def to_html(df):
            html = "<table style='width:100%; border-collapse: collapse;'>"
            html += "<thead><tr>" + "".join([
                f"<th style='border:1px solid #ccc; padding:4px'>{col}</th>"
                for col in df.columns[:-1]
            ]) + "</tr></thead><tbody>"
            for _, row in df.iterrows():
                style = "background-color:#f0f0f0;" if row["IsFirst"] else ""
                html += "<tr>" + "".join([
                    f"<td style='border:1px solid #ccc; padding:4px; {style}'>{row[col]}</td>"
                    for col in df.columns[:-1]
                ]) + "</tr>"
            html += "</tbody></table>"
            return html

        st.markdown(to_html(df), unsafe_allow_html=True)

        # ✅ 匯出 Excel 並加樣式
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            clean_df = df.drop(columns=["IsFirst"])
            clean_df.to_excel(tmp.name, index=False)

            wb = openpyxl.load_workbook(tmp.name)
            ws = wb.active

            gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            bold_font = Font(bold=True)

            for i, is_first in enumerate(df["IsFirst"], start=2):  # Excel 從第2列開始（第1列是標題）
                if is_first:
                    ws[f"B{i}"].fill = gray_fill
                    ws[f"B{i}"].font = bold_font

            wb.save(tmp.name)

            with open(tmp.name, "rb") as f:
                st.download_button("⬇️ 下載 Excel", f, file_name="authors.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.warning("⚠️ 沒有找到任何作者資料。")
